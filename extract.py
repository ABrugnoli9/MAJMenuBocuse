# -*- coding: utf-8 -*-
"""
Extraction hebdomadaire des "Cost carte" (1er onglet uniquement) -> snapshot JSON + page HTML.

- Lit UNIQUEMENT le premier onglet ("Récap. cost carte") de chaque fichier maison.
- Ignore le reste du classeur (BDD, fiches techniques...) -> rapide et léger.
- Sauvegarde un instantané de la semaine dans data/snapshot-AAAA-Www.json
- Compare au dernier instantané précédent (evolution du coût de revient).
- Génère index.html (page unique, sélecteur de maison + récap chef).

Usage :  python extract.py
"""
import os, re, glob, json, datetime, sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl manquant : pip install openpyxl")

# --- Configuration ---------------------------------------------------------
DOSSIER_CARTES = r"P:\SIEGE\Controle de Gestion\Fiches techniques\FT - 2025-#\0. MAISONS - Cartes\2026\3. Carte ÉTÉ - 2026"
ICI = os.path.dirname(os.path.abspath(__file__))
DOSSIER_DATA = os.path.join(ICI, "data")

# Colonnes (index 0-based) dans l'onglet Récap
COL_CATEGORIE = 1
COL_STATUT    = 2
COL_NOM       = 3
COL_NOM_COURT = 4
COL_COUT      = 11
COL_PV        = 12
COL_RATIO     = 13   # Ratio théorique
COL_PV_CIBLE  = 14   # PV TTC (cible 26%)
COL_ECART     = 15   # Ecart
LIGNE_ENTETE  = 4   # ligne Excel des en-têtes (1-based)


def num(v):
    """Convertit en float propre ou None."""
    if v is None or v == "":
        return None
    try:
        return round(float(v), 4)
    except (TypeError, ValueError):
        return None


def nom_maison(chemin):
    """'1. Cost carte - NORD (E. 2026).xlsm' -> 'NORD'."""
    base = os.path.basename(chemin)
    m = re.search(r"Cost carte\s*-\s*(.+?)\s*\(", base)
    return m.group(1).strip() if m else base


def lire_maison(chemin):
    """Lit le 1er onglet et renvoie {maison, saison, plats:[...]}"""
    wb = openpyxl.load_workbook(chemin, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]              # premier onglet uniquement
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    saison = ""
    if len(rows) > 1 and rows[1] and len(rows[1]) > 2 and rows[1][2]:
        saison = str(rows[1][2]).strip()

    plats = []
    for row in rows[LIGNE_ENTETE:]:        # après la ligne d'en-tête
        if not row:
            continue
        nom = row[COL_NOM] if len(row) > COL_NOM else None
        if not nom or str(nom).strip() in ("", "0"):
            continue                        # ligne vide / réserve
        cout = num(row[COL_COUT]) if len(row) > COL_COUT else None
        if cout is None:
            continue
        plats.append({
            "categorie": str(row[COL_CATEGORIE]).strip() if len(row) > COL_CATEGORIE and row[COL_CATEGORIE] else "",
            "statut":    str(row[COL_STATUT]).strip() if len(row) > COL_STATUT and row[COL_STATUT] else "",
            "nom":       str(nom).strip(),
            "nom_court": str(row[COL_NOM_COURT]).strip() if len(row) > COL_NOM_COURT and row[COL_NOM_COURT] else str(nom).strip(),
            "cout":      cout,
            "pv":        num(row[COL_PV]) if len(row) > COL_PV else None,
            "ratio":     num(row[COL_RATIO]) if len(row) > COL_RATIO else None,
            "pv_cible":  num(row[COL_PV_CIBLE]) if len(row) > COL_PV_CIBLE else None,
            "ecart":     num(row[COL_ECART]) if len(row) > COL_ECART else None,
        })
    return {"maison": nom_maison(chemin), "saison": saison, "plats": plats}


def collecter():
    """Balaye tous les fichiers 'Cost carte - *' (hors BDD)."""
    motif = os.path.join(DOSSIER_CARTES, "*Cost carte*-*.xlsm")
    fichiers = [f for f in glob.glob(motif) if "BDD" not in os.path.basename(f)]
    fichiers.sort()
    maisons = []
    for f in fichiers:
        try:
            d = lire_maison(f)
            print(f"  OK  {d['maison']:<12} {len(d['plats'])} plats")
            maisons.append(d)
        except Exception as e:
            print(f"  ERR {os.path.basename(f)} : {e}")
    return maisons


def semaine_iso(dt=None):
    dt = dt or datetime.date.today()
    y, w, _ = dt.isocalendar()
    return f"{y}-W{w:02d}"


def snapshot_precedent(semaine_actuelle):
    """Renvoie le contenu du dernier snapshot AVANT la semaine actuelle."""
    fichiers = sorted(glob.glob(os.path.join(DOSSIER_DATA, "snapshot-*.json")))
    prec = None
    for f in fichiers:
        w = os.path.basename(f).replace("snapshot-", "").replace(".json", "")
        if w < semaine_actuelle:
            prec = f
    if not prec:
        return None
    with open(prec, encoding="utf-8") as fh:
        return json.load(fh)


def index_plats(snap):
    """Dict {maison: {nom_plat: cout}} depuis un snapshot."""
    idx = {}
    for m in snap.get("maisons", []):
        idx[m["maison"]] = {p["nom"]: p["cout"] for p in m["plats"]}
    return idx


def main():
    os.makedirs(DOSSIER_DATA, exist_ok=True)
    semaine = semaine_iso()
    print(f"Semaine {semaine} — lecture des cartes...")
    maisons = collecter()
    if not maisons:
        sys.exit("Aucune carte lue.")

    snap = {
        "genere_le": datetime.datetime.now().isoformat(timespec="seconds"),
        "semaine": semaine,
        "maisons": maisons,
    }

    # Comparaison avec la semaine précédente
    prec = snapshot_precedent(semaine)
    idx_prec = index_plats(prec) if prec else {}
    for m in maisons:
        anc = idx_prec.get(m["maison"], {})
        for p in m["plats"]:
            ancien = anc.get(p["nom"])
            p["cout_prec"] = ancien
            if ancien is None or p["cout"] is None:
                p["delta"] = None
                p["delta_pct"] = None
            else:
                p["delta"] = round(p["cout"] - ancien, 4)
                p["delta_pct"] = round((p["cout"] - ancien) / ancien * 100, 2) if ancien else None

    # Sauvegarde snapshot de la semaine
    chemin_snap = os.path.join(DOSSIER_DATA, f"snapshot-{semaine}.json")
    with open(chemin_snap, "w", encoding="utf-8") as fh:
        json.dump(snap, fh, ensure_ascii=False, indent=2)
    print(f"Snapshot -> {chemin_snap}")

    # Génération HTML (page principale, nom fixe -> lien constant)
    from generer_html import generer
    html = generer(snap, prec)
    chemin_html = os.path.join(ICI, "index.html")
    with open(chemin_html, "w", encoding="utf-8") as fh:
        fh.write(html)
    print(f"Page   -> {chemin_html}")

    # Archive de la semaine (copie autonome) + index des archives
    dossier_arch = os.path.join(ICI, "archives")
    os.makedirs(dossier_arch, exist_ok=True)
    chemin_arch = os.path.join(dossier_arch, f"cost-carte-{semaine}.html")
    # La copie est DANS archives/ : le lien vers la liste doit rester dans ce dossier
    html_arch = html.replace('href="archives/index.html"', 'href="index.html"')
    with open(chemin_arch, "w", encoding="utf-8") as fh:
        fh.write(html_arch)
    print(f"Archive-> {chemin_arch}")
    generer_index_archives(dossier_arch)
    print("Terminé.")


def generer_index_archives(dossier_arch):
    """Construit archives/index.html : la liste de toutes les semaines archivées."""
    fichiers = sorted(glob.glob(os.path.join(dossier_arch, "cost-carte-*.html")), reverse=True)
    liens = ""
    for f in fichiers:
        nom = os.path.basename(f)
        sem = nom.replace("cost-carte-", "").replace(".html", "")
        liens += f'<li><a href="{nom}">Semaine {sem}</a></li>\n'
    if not liens:
        liens = "<li><i>Aucune archive pour le moment.</i></li>"
    html = f"""<!DOCTYPE html><html lang="fr"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Archives — Cost Carte</title>
<style>
 body{{font-family:-apple-system,Segoe UI,Roboto,Arial,sans-serif;background:#f6f7f9;color:#1c2430;margin:0}}
 header{{background:#8a1538;color:#fff;padding:18px 22px}} header h1{{margin:0;font-size:20px}}
 .wrap{{max-width:700px;margin:0 auto;padding:24px 22px}}
 ul{{list-style:none;padding:0}} li{{margin:0}}
 li a{{display:block;padding:13px 16px;background:#fff;border:1px solid #e6e9ed;border-radius:10px;margin-bottom:10px;
      text-decoration:none;color:#8a1538;font-weight:600;font-size:16px}}
 li a:hover{{background:#faf7f8}}
 a.retour{{color:#8a1538}}
</style></head><body>
<header><h1>📅 Archives — Cost Carte Maisons Bocuse</h1></header>
<div class="wrap">
 <p><a class="retour" href="../index.html">← Retour à la semaine en cours</a></p>
 <ul>
{liens}</ul>
</div></body></html>"""
    with open(os.path.join(dossier_arch, "index.html"), "w", encoding="utf-8") as fh:
        fh.write(html)


if __name__ == "__main__":
    main()
